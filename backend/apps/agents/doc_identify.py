"""Automatic document-type identification.

Strategy (cheap → expensive):

1. Filename heuristics — strong signals like 'CIBIL', 'BANK_STATEMENT', 'GSTIN',
   'ITR', 'AA-' (Account Aggregator) are highly reliable.
2. Content peek — open the file and extract ~1-2kb of text from:
   - PDF: first page via pypdf
   - XLSX: cell values from first sheet
   - JSON/CSV/TXT: first bytes
3. Keyword scoring against a map of doc types → indicative phrases.

The function returns a ``(doc_type, confidence, evidence)`` tuple. ``evidence`` is
a short human-readable snippet we can show in the UI so reviewers can sanity-
check the auto-classification at a glance.
"""
from __future__ import annotations

import csv
import io
import logging
import re
from pathlib import Path
from typing import Any

from apps.cases.models import DocumentType

logger = logging.getLogger(__name__)


FILENAME_HINTS: list[tuple[str, str]] = [
    ("cibil", DocumentType.CIBIL),
    ("experian", DocumentType.CIBIL),
    ("crif", DocumentType.CIBIL),
    ("equifax", DocumentType.CIBIL),
    ("credit_report", DocumentType.CIBIL),
    ("bank-statement", DocumentType.BANK_STATEMENT),
    ("bank_statement", DocumentType.BANK_STATEMENT),
    ("bankstatement", DocumentType.BANK_STATEMENT),
    ("statement", DocumentType.BANK_STATEMENT),
    ("passbook", DocumentType.BANK_STATEMENT),
    ("gst_certificate", DocumentType.GST_CERTIFICATE),
    ("gst-certificate", DocumentType.GST_CERTIFICATE),
    ("gstcertificate", DocumentType.GST_CERTIFICATE),
    ("gst-reg", DocumentType.GST_CERTIFICATE),
    ("gstin", DocumentType.GSTIN_FILINGS),
    ("gst_filing", DocumentType.GSTIN_FILINGS),
    ("gst-filing", DocumentType.GSTIN_FILINGS),
    ("gstr", DocumentType.GSTIN_FILINGS),
    ("ledger", DocumentType.LEDGER),
    ("tally", DocumentType.LEDGER),
    ("itr", DocumentType.ITR),
    ("26as", DocumentType.ITR),
    ("aa", DocumentType.BANK_AC_AUTH),
    ("account_aggregator", DocumentType.BANK_AC_AUTH),
    ("aa-", DocumentType.BANK_AC_AUTH),
]

CONTENT_KEYWORDS: dict[str, list[str]] = {
    DocumentType.CIBIL: [
        r"\bCIBIL\b",
        r"credit\s+information\s+report",
        r"cibil\s+trans?\s?union",
        r"credit\s+score",
        r"bureau\s+score",
        r"\bTransUnion\b",
        r"Experian",
        r"CRIF\s+High\s?Mark",
        r"Equifax",
    ],
    DocumentType.BANK_STATEMENT: [
        r"bank\s+statement",
        r"statement\s+of\s+account",
        r"account\s+number",
        r"closing\s+balance",
        r"opening\s+balance",
        r"IFSC",
        r"UPI",
        r"NEFT",
        r"RTGS",
        r"transaction\s+date",
    ],
    DocumentType.GST_CERTIFICATE: [
        r"certificate\s+of\s+registration",
        r"GSTIN",
        r"goods\s+and\s+services\s+tax",
        r"form\s+gst\s+reg-06",
        r"legal\s+name\s+of\s+business",
    ],
    DocumentType.GSTIN_FILINGS: [
        r"gstr[-\s]?1",
        r"gstr[-\s]?3b",
        r"sales?\s+return",
        r"outward\s+supplies",
        r"monthly\s+return",
        r"fy\s?20\d{2}",
    ],
    DocumentType.ITR: [
        r"income\s+tax\s+return",
        r"itr[-\s]?[1-7]",
        r"assessing\s+officer",
        r"gross\s+total\s+income",
        r"assessment\s+year",
    ],
    DocumentType.LEDGER: [
        r"ledger\s+account",
        r"debit.*credit",
        r"opening\s+balance.*closing\s+balance",
        r"journal\s+entries",
        r"trial\s+balance",
    ],
    DocumentType.BANK_AC_AUTH: [
        r"account\s+aggregator",
        r"consent\s+handle",
        r"reBit",
        r"sahamati",
        r"fiu[-\s]",
        r"consent\s+detail",
    ],
}


def _extract_text_pdf(path: Path, max_chars: int = 4000) -> str:
    try:
        import pypdf  # type: ignore[import-not-found]
    except Exception:
        return ""
    try:
        reader = pypdf.PdfReader(str(path))
        parts: list[str] = []
        for page in reader.pages[:2]:
            text = page.extract_text() or ""
            parts.append(text)
            if sum(len(p) for p in parts) >= max_chars:
                break
        return "\n".join(parts)[:max_chars]
    except Exception as exc:  # noqa: BLE001
        logger.debug("pdf text extract failed: %s", exc)
        return ""


def _extract_text_xlsx(path: Path, max_chars: int = 4000) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except Exception:
        return ""
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.worksheets[:2]:
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i > 60:
                    break
                for cell in row[:12]:
                    if cell is None:
                        continue
                    parts.append(str(cell))
                    if sum(len(p) for p in parts) >= max_chars:
                        return " ".join(parts)[:max_chars]
        return " ".join(parts)[:max_chars]
    except Exception as exc:  # noqa: BLE001
        logger.debug("xlsx text extract failed: %s", exc)
        return ""


def _extract_text_csv(path: Path, max_chars: int = 4000) -> str:
    try:
        with path.open(newline="", errors="replace") as fh:
            reader = csv.reader(fh)
            parts: list[str] = []
            for i, row in enumerate(reader):
                if i > 100:
                    break
                parts.append(",".join(row))
                if sum(len(p) for p in parts) >= max_chars:
                    break
            return "\n".join(parts)[:max_chars]
    except Exception:
        return ""


def _extract_text_generic(path: Path, max_chars: int = 4000) -> str:
    try:
        with path.open("rb") as fh:
            raw = fh.read(max_chars * 2)
        return raw.decode("utf-8", errors="ignore")[:max_chars]
    except Exception:
        return ""


def _peek(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return _extract_text_pdf(path)
    if suffix in {".xlsx", ".xls"}:
        return _extract_text_xlsx(path)
    if suffix == ".csv":
        return _extract_text_csv(path)
    if suffix in {".json", ".txt", ".xml"}:
        return _extract_text_generic(path)
    return _extract_text_generic(path)


def _score_content(text: str) -> dict[str, tuple[float, str]]:
    if not text:
        return {}
    scores: dict[str, tuple[float, str]] = {}
    lowered = text.lower()
    for doc_type, patterns in CONTENT_KEYWORDS.items():
        hits = 0
        evidence = ""
        for pat in patterns:
            match = re.search(pat, lowered, re.IGNORECASE)
            if match:
                hits += 1
                if not evidence:
                    start = max(0, match.start() - 20)
                    evidence = text[start : match.end() + 40].strip().replace("\n", " ")
        if hits:
            norm = min(1.0, hits / 3.0)
            scores[doc_type] = (norm, evidence)
    return scores


def _filename_score(filename: str) -> tuple[str | None, float]:
    lower = filename.lower()
    for needle, kind in FILENAME_HINTS:
        if needle in lower:
            return kind, 0.7
    return None, 0.0


def identify(path: Path, filename: str | None = None) -> dict[str, Any]:
    """Classify a document. Returns {doc_type, confidence, source, evidence}."""
    name = filename or path.name
    fn_type, fn_conf = _filename_score(name)

    text = _peek(path)
    content_scores = _score_content(text)

    best_doc = fn_type
    best_conf = fn_conf
    best_source = "filename" if fn_type else "unknown"
    best_evidence = name if fn_type else ""

    for doc_type, (conf, evidence) in content_scores.items():
        combined = conf + (0.15 if doc_type == fn_type else 0.0)
        if combined > best_conf:
            best_doc = doc_type
            best_conf = min(1.0, combined)
            best_source = "content" if doc_type != fn_type else "filename+content"
            best_evidence = evidence[:180]

    if not best_doc:
        best_doc = DocumentType.OTHER
        best_conf = 0.3
        best_source = "fallback"

    return {
        "doc_type": best_doc,
        "confidence": round(best_conf, 2),
        "source": best_source,
        "evidence": best_evidence,
    }
